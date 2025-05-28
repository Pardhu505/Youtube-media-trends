import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import os
import logging

# Configure basic logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Determine the base directory of the Flask application (flask_youtube_app)
# __file__ in excel_utils.py is flask_youtube_app/excel_utils.py
# os.path.dirname(__file__) is flask_youtube_app
BASE_APP_DIR = os.path.dirname(os.path.abspath(__file__))
STATIC_DIR = os.path.join(BASE_APP_DIR, 'static')

def save_to_excel(data, output_filename, format_duration_func):
    """
    Saves video data to an Excel file, including thumbnails.
    """
    if not data:
        logging.warning("No data provided to save_to_excel.")
        return None

    final_export_data = []
    for item_orig in data:
        item = item_orig.copy() # Work with a copy
        
        # Ensure all expected numeric fields are present and are numbers, default to 0 if not
        views = item.get('views', 0)
        likes = item.get('likes', 0)
        comments = item.get('comments', 0)
        duration_seconds = item.get('duration_seconds')

        try:
            views = int(views) if views is not None else 0
        except (ValueError, TypeError):
            views = 0
        try:
            likes = int(likes) if likes is not None else 0
        except (ValueError, TypeError):
            likes = 0
        try:
            comments = int(comments) if comments is not None else 0
        except (ValueError, TypeError):
            comments = 0
        
        formatted_duration = "N/A"
        if duration_seconds is not None:
            try:
                formatted_duration = format_duration_func(int(duration_seconds))
            except (ValueError, TypeError):
                logging.warning(f"Could not format duration for seconds: {duration_seconds}")
                formatted_duration = "N/A" # Fallback if formatting fails
        
        excel_item = {
            'Title': item.get('title', 'N/A'),
            'URL': item.get('url', 'N/A'),
            'Channel': item.get('channel_name', 'N/A'),
            'Views': views,
            'Duration': formatted_duration,
            'Likes': likes,
            'Comments': comments,
            'Date': item.get('date', 'N/A') 
        }
        final_export_data.append(excel_item)
    
    df = pd.DataFrame(final_export_data)
    
    # Define the desired column order for the Excel sheet
    # This order should match the headers you want in the Excel file.
    # The thumbnail images will be added to an additional column, not part of this DataFrame directly.
    desired_columns = ['Title', 'URL', 'Channel', 'Views', 'Duration', 'Likes', 'Comments', 'Date']
    df = df[desired_columns]

    try:
        df.to_excel(output_filename, index=False, engine='openpyxl')
        wb = openpyxl.load_workbook(output_filename)
        ws = wb.active

        # Set column widths for data columns (A-H for 'Title' through 'Date')
        ws.column_dimensions['A'].width = 50  # Title
        ws.column_dimensions['B'].width = 40  # URL
        ws.column_dimensions['C'].width = 25  # Channel
        ws.column_dimensions['D'].width = 10  # Views
        ws.column_dimensions['E'].width = 15  # Duration
        ws.column_dimensions['F'].width = 10  # Likes
        ws.column_dimensions['G'].width = 10  # Comments
        ws.column_dimensions['H'].width = 15  # Date
        
        # Thumbnail Image Column settings (Column I)
        thumbnail_col_letter = 'I'
        ws.column_dimensions[thumbnail_col_letter].width = 20 
        ws[f'{thumbnail_col_letter}1'] = "Thumbnail" # Header for thumbnail column

        for idx, video_data_item in enumerate(data, start=2): # enumerate from original data for thumbnail paths
            thumb_relative_path = video_data_item.get('thumbnail') # This is like 'thumbnails/image.jpg'
            
            if thumb_relative_path and isinstance(thumb_relative_path, str) and thumb_relative_path.lower() != 'n/a':
                thumb_full_path = os.path.join(STATIC_DIR, thumb_relative_path)
                
                if os.path.exists(thumb_full_path):
                    try:
                        img = XLImage(thumb_full_path)
                        img.width, img.height = 120, 67.5 # Aspect ratio 16:9
                        ws.row_dimensions[idx].height = 55 # Approx height for image + padding
                        ws.add_image(img, f"{thumbnail_col_letter}{idx}")
                    except UnidentifiedImageError:
                        logging.error(f"Cannot identify image (it may be corrupted or not a supported format): {thumb_full_path}")
                        ws[f"{thumbnail_col_letter}{idx}"] = "Error: Bad Image"
                    except Exception as e:
                        logging.error(f"Error adding image {thumb_full_path} to Excel: {e}")
                        ws[f"{thumbnail_col_letter}{idx}"] = "Error: Image"
                else:
                    logging.warning(f"Thumbnail file not found: {thumb_full_path}")
                    ws[f"{thumbnail_col_letter}{idx}"] = "Not Found"
            else:
                ws[f"{thumbnail_col_letter}{idx}"] = "N/A" # No thumbnail path provided
        
        wb.save(output_filename)
        logging.info(f"Excel file '{output_filename}' created successfully with thumbnails.")
        return output_filename
    except Exception as e:
        logging.error(f"Error creating Excel file '{output_filename}': {e}", exc_info=True)
        return None
